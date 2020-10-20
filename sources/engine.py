import numpy as np
from multiprocessing import Process, Queue
from PIL import Image
import cv2
from .common.constants import *
from skimage import draw
from skimage.transform import resize
from openpyxl import load_workbook
import os
import tensorflow as tf
from tensorflow import keras
from .model_loader import get_model
import pickle
import json
from pathlib import Path

# To limit loop rate
from pygame.time import Clock

class Engine(Process):
    """
    Main process that calculates all the necessary computations
    """
    # If the image is not updated, check if self._updated is switched to True
    def __init__(self, to_EngineQ:Queue, to_ConsoleQ:Queue,
                 imageQ:Queue, eventQ:Queue, etcQ:Queue):
        super().__init__(daemon=True)
        # Initial image and mask
        self.image = np.zeros((300,300,3), dtype=np.uint8)
        # Queues
        self._to_EngineQ = to_EngineQ
        self._to_ConsoleQ = to_ConsoleQ
        self._imageQ = imageQ
        self._eventQ = eventQ
        self._etcQ = etcQ
        # Modes about sending images to Viewer
        self._mask_mode = False
        self._updated = True
        # Do only one thing at a time - Do not make multiple modes
        self._mode = None
        # Initial membrane and cell color (NOT MASK COLORS)
        self.mem_color = MEMBRANE
        self.cell_color = CELL
        # Default mask ratio
        self._mask_ratio = 0.5
        # Box layers
        self._box_layers = []
        self._box_start_pos = None
        # Clipped mode
        self._clipped_mode = False
        self._clipped_masks = []
        self._clipped_imgs = []
        # (Color_of_layer(R,G,B), (xx, yy))
        # (xx, yy) = np.nonzero(array)
        self._layers = []
        self._cell_layers = []
        self._cell_counts = []
        # To viewer (x, y)
        self._cell_pos = []

        self._always_on_layers = []
        self._is_drawing = False
        self._line_start_pos = None
        self._show_box = True
        # Modes related to filling
        # Ratio = (micrometer / pixel)**2  -> Because it's area ratio
        self._mp_ratio = DEFAULT_MP_RATIO
        self._mp_ratio_pixel = DEFAULT_MP_PIXEL_ORIGINAL * DEFAULT_RESIZED_RATIO
        self._mp_ratio_pixel_original = DEFAULT_MP_PIXEL_ORIGINAL
        self._mp_ratio_micrometer = DEFAULT_MP_MICRO
        self._resized_ratio = DEFAULT_RESIZED_RATIO
        # Data to save
        self._data = []

    @property
    def image(self):
        """
        This is a base image. Do not directly modify this
        """
        return self._image.copy()

    @image.setter
    def image(self, image:np.array):
        """
        Must be a shape of (width, height, 3)
        """
        if len(image.shape) != 3 and image.shape[2] != 3:
            raise TypeError('Inappropriate shape of image')
        self._image = image.astype(np.uint8)
        self._shape = self._image.shape
        # self.set_empty_mask()
        self._updated = True

    @property
    def shape(self):
        """
        Do not implement shape.setter
        Shape is dependent to image and should only be setted with image
        """
        return self._shape

    # @property
    # def mask(self):
    #     """
    #     This is the mask on which engine computes
    #     """
    #     return self._mask.copy()

    # @mask.setter
    # def mask(self, mask:np.array):
    #     """
    #     Must be a shape of (width, height, 3) and same as current image
    #     """
    #     if mask.shape != self.shape:
    #         raise TypeError('Inappropriate shape of mask')
    #     self._mask = mask.astype(np.uint8)

    @property
    def cell_color(self):
        return self._cell_color

    @cell_color.setter
    def cell_color(self, color):
        if len(color) != 3 :
            raise TypeError('Wrong color given')
        self._cell_color = color

    @property
    def mem_color(self):
        return self._mem_color

    @mem_color.setter
    def mem_color(self, color):
        if len(color) != 3 :
            raise TypeError('Wrong color given')
        self._mem_color = color

    @property
    def mask_mode(self):
        return self._mask_mode

    @mask_mode.setter
    def mask_mode(self, mask_mode:bool):
        self._mask_mode = mask_mode
        self._updated = True

    @property
    def mode(self):
        return self._mode

    @mode.setter
    def mode(self, mode):
        if self._mode == MODE_DRAW_MEM:
            self.draw_stop()
        elif self._mode == MODE_FILL_MP_RATIO and self._is_drawing:
            self.fill_ratio_cancel()
        elif self._mode == MODE_DRAW_BOX:
            self.draw_box_stop()
        self._mode = mode
        self._updated = True

    def reset(self):
        self._layers = []
        self._cell_layers = []
        self._cell_counts = []
        self._cell_pos = []
        self._box_layers = []
        self._always_on_layers = []
        self._clipped_imgs = []
        self._clipped_masks = []
        self._data = []
        self._is_drawing = False
        self._line_start_pos = None
        self._box_start_pos = None
        self.mode = None
        self._mask_mode = False
        self._updated = True

    def load_image(self, path:str):
        im = Image.open(path)
        self._original_size = im.size
        self._original_image = np.asarray(im)
        o_width, o_height = im.size
        if o_width*3 >= o_height*4:
            # if wider than taller, fix width to 1200
            new_width = 1200
            new_height = o_height*1200//o_width
        else:
            # if taller than wider, fix height to 900
            new_height = 900
            new_width = o_width*900//o_height
        self._resized_ratio = new_width / o_width
        self.image = np.asarray(im.resize((new_width, new_height))).swapaxes(0,1)
        self.reset()
        self._updated = True


    def change_mask_ratio(self, ratio:float):
        self._mask_ratio = ratio / 100
        self.mode = None
        msg = f'Ratio set to {self._mask_ratio}'
        self._to_ConsoleQ.put({MESSAGE_BOX:msg})
        self._updated = True
    
    def put_image(self):
        tmp_image = self.image
        if self._mask_mode:
            for c, m in self._cell_layers:
                xx, yy = m
                tmp_image[xx, yy] = c
        for c, m in self._always_on_layers:
            xx, yy = m
            tmp_image[xx, yy] = c
        self._imageQ.put(tmp_image)

    def put_mode(self):
        if self.mode != None:
            self._to_ConsoleQ.put({self.mode:None})
        else:
            self._to_ConsoleQ.put({MODE_NONE:None})

        original_ratio = self._mp_ratio * (self._resized_ratio**2)
        self._to_ConsoleQ.put({FILL_MP_RATIO:original_ratio})
        self._to_ConsoleQ.put({FILL_PIXEL:self._mp_ratio_pixel_original})

    def put_ratio_list(self):
        self._mp_ratio = (self._mp_ratio_micrometer/self._mp_ratio_pixel)**2
        area_list = np.multiply(self._cell_counts, self._mp_ratio).tolist()
        self._to_ConsoleQ.put({FILL_LIST:area_list})
        self._etcQ.put({
            POS_LIST:zip(self._cell_pos, area_list)
            })

    def draw_box_start(self, pos):
        """
        Make a new layer and draw initial point (Red dot)
        """
        new_layer = np.zeros((self.shape[0],self.shape[1]),
                             dtype=np.bool)
        color = BOX_START
        x, y = pos
        new_layer[:, y] = True
        new_layer[x,:] = True
        self._always_on_layers.append((color, np.nonzero(new_layer)))
        self._box_start_pos = pos
        self._is_drawing = True
        self._updated = True

    def draw_box_stop(self):
        """
        When drawing box is interrupted.
        """
        self._box_start_pos = None
        if self._is_drawing:
            self._always_on_layers.pop()
        self._is_drawing = False
        self._etcQ.put({CROSS_CURSOR_OFF:None})
        self._updated = True


    def draw_box_end(self, pos):
        """
        Draw Box
        """
        self._always_on_layers.pop()
        # If user clicked on the same axis (no width or height), skip
        if not np.any(np.equal(self._box_start_pos, pos)):
            color = np.random.randint(0,255,3)
            x0, y0 = self._box_start_pos
            x1, y1 = pos
            r0, c0 = min(x0, x1), min(y0, y1)
            r1, c1 = max(x0, x1), max(y0, y1)
            clipped_image =  self.image[r0:r1,c0:c1]
            casted_input = resize(clipped_image, (200,200), preserve_range=True,
                            anti_aliasing=True)[np.newaxis,:].astype(np.float32)
            raw_output = self._mask_model(casted_input).numpy()[0]
            prob_mask = resize(raw_output, clipped_image.shape[:2], 
                            preserve_range=True, anti_aliasing=True)
            mask = (prob_mask > self._mask_ratio)
            if not np.any(mask):
                # If there's no pixels detected as cell, skip
                self._to_ConsoleQ.put({
                    MESSAGE_BOX:'No pixels detected as a cell'
                })
            else:
                xx, yy = np.nonzero(mask)
                np.add(xx, r0, out=xx)
                np.add(yy, c0, out=yy)
                self._cell_layers.append((color,(xx, yy)))
                self._cell_counts.append(len(xx))

                # To viewer (position of cell)
                avgx = np.average(xx)
                avgy = np.average(yy)
                self._cell_pos.append((avgx, avgy))

                datum = {}
                datum['box'] = [[r0,c0],[r1,c1]]
                datum['mask'] = [xx.tolist(), yy.tolist()]
                datum['size'] = self.shape[:2]
                self._data.append(datum)
        self._mask_mode = True
        self._box_start_pos = None
        self._is_drawing = False
        self.mode = None
        self._etcQ.put({CROSS_CURSOR_OFF:None})
        self._updated = True

    def fill_ratio_start(self, pos):
        new_layer = np.zeros((self.shape[0],self.shape[1]),
                             dtype=np.bool)
        color = LINE_START
        x, y = pos
        new_layer[x-2:x+3, y-2:y+3] = True
        self._always_on_layers.append((color, np.nonzero(new_layer)))
        self._mp_ratio_start_pos = pos
        self._is_drawing = True
        self._updated = True

    def fill_ratio_end(self, pos):
        pixel_dist = np.sqrt(np.sum(np.subtract(self._mp_ratio_start_pos,pos)**2))
        # Do not update if clicked the same position
        if pixel_dist != 0 :
            self._mp_ratio_pixel = pixel_dist
            self._mp_ratio_pixel_original = pixel_dist/self._resized_ratio
        self._always_on_layers.pop()
        self._is_drawing = False
        self._mp_ratio_start_pos = None
        self.mode = None
        self._updated = True
    
    def fill_ratio_cancel(self):
        self._mp_ratio_start_pos = None
        self._is_drawing = False
        self._always_on_layers = []
        self._updated = True

    def fill_delete(self, indices):
        if len(self._cell_counts) > 0:
            for idx in indices:
                self._cell_layers.pop(idx)
                self._cell_counts.pop(idx)
                self._cell_pos.pop(idx)
                self._data.pop(idx)
            self._updated = True

    def fill_save(self, excel_dir, image_name, image_folder):
        try :
            wb = load_workbook(excel_dir)
        except :
            self._to_ConsoleQ.put({MESSAGE_BOX:'Cannot open Workbook!'})
            return
        ws = wb.worksheets[0]
        row, col = 1, 1
        while ws.cell(row, col).value != None:
            col += 1
        for area in self._cell_counts:
            ws.cell(row, col).value = area * self._mp_ratio
            ws.cell(row, col+1).value = image_name
            row += 1
        try:
            wb.save(excel_dir)
        except:
            self._to_ConsoleQ.put({MESSAGE_BOX:'Failed to Save'})
            return
        else:
            self._to_ConsoleQ.put({MESSAGE_BOX:'Excel file Saved Successfully.'\
                '\nDon\'t forget to check.'})
        # Saving the data
        image_folder = Path(image_folder)
        save_folder = image_folder / 'save'
        if not save_folder.exists():
            save_folder.mkdir()

        start_num = len(os.listdir(str(save_folder)))
        data_name = str(start_num) + '.json'

        filename_data = save_folder / data_name

        for datum in self._data:
            datum['image'] = image_name

        with open(str(filename_data), 'w') as f:
            json.dump(self._data, f, indent=4)

    def save_screenshot(self, image_path):
        resized_masks = []
        original_copy = self._original_image.copy()
        area_list = np.multiply(self._cell_counts, self._mp_ratio).tolist()
        for i in range(len(self._cell_layers)):
            dummy_layer = np.zeros(self.shape[:2],dtype=np.uint8)
            c, m = self._cell_layers[i]
            xx, yy = m
            dummy_layer[xx,yy] = 1
            dummy_layer = cv2.resize(dummy_layer.swapaxes(0,1),
                                    dsize=self._original_size,
                                    interpolation=cv2.INTER_LINEAR)
            xx, yy = np.nonzero(dummy_layer)
            original_copy[xx,yy] = c

            new_pos = np.divide(self._cell_pos[i],self._resized_ratio).astype(np.int)
            r_start = max(new_pos[1]-10,0)
            c_start = max(new_pos[0]-30,0)

            original_copy[r_start:r_start+20,
                          c_start:c_start+80]=255

            cv2.putText(
                img=original_copy,
                text=f'{area_list[i]:.2f}',
                org=tuple(np.clip(new_pos-[30,-5],0,None)),
                fontFace=cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=0.5,
                thickness=1,
                color=(255,0,0),
            )
        try:
            Image.fromarray(original_copy).save(image_path)
        except:
            self._to_ConsoleQ.put({MESSAGE_BOX:'Failed to Save screenshot'})
            return
        self._to_ConsoleQ.put({MESSAGE_BOX:'Saved screenshot'})

    def run(self):
        mainloop = True
        self._clock = Clock()
        self._mask_model = get_model('hr_5_3_0')
        while mainloop:
            self._clock.tick(60)
            if not self._to_EngineQ.empty():
                q = self._to_EngineQ.get()
                for k, v in q.items():
                    if k == TERMINATE:
                        mainloop = False
                    # Loading & Showing modes
                    elif k == NEWIMAGE:
                        self.load_image(v)
                    # elif k == NEWMASK:
                    #     self.set_new_mask()
                    elif k == MODE_IMAGE:
                        self.mask_mode = False
                    elif k == MODE_MASK:
                        self.mask_mode = True
                    # Box Drawing
                    elif k == DRAW_BOX:
                        # NOTE: Same thing should be in K_B (keyboard shortcut)
                        self.mode = MODE_DRAW_BOX
                        self._etcQ.put({CROSS_CURSOR_ON:None})
                        self._updated = True
                    elif k == SET_RATIO:
                        self.change_mask_ratio(v)
                    elif k == MODE_CANCEL_BOX:
                        self.mode = None
                    #Counting modes
                    elif k == FILL_MP_RATIO:
                        self.mode = MODE_FILL_MP_RATIO
                        self._updated = True
                    elif k == FILL_DELETE:
                        self.fill_delete(v)
                    elif k == FILL_SAVE:
                        self.fill_save(*v)
                    elif k == FILL_MICRO:
                        self._mp_ratio_micrometer = v
                        self._updated = True
                    elif k == FILL_PIXEL:
                        self._mp_ratio_pixel_original = v
                        self._mp_ratio_pixel = v*self._resized_ratio
                        self._updated = True
                    elif k == FILL_SCREENSHOT:
                        self.save_screenshot(v)

            if not self._eventQ.empty():
                q = self._eventQ.get()
                for k,v in q.items():
                    if k == MOUSEDOWN:
                        # v : mouse pos which came from Viewer
                        # Box mode
                        if self.mode == MODE_DRAW_BOX:
                            if not self._is_drawing:
                                self.draw_box_start(v)
                            else:
                                self.draw_box_end(v)
                        # # Counting mode
                        elif self.mode == MODE_FILL_MP_RATIO:
                            if not self._is_drawing:
                                self.fill_ratio_start(v)
                            else :
                                self.fill_ratio_end(v)
                    elif k == MOUSEDOWN_RIGHT:
                        pass
                    elif k == MOUSEUP:
                        pass
                    elif k == MOUSEPOS:
                        pass
                    # Keyboard events
                    elif k == K_Z:
                        self.fill_delete([-1])
                    elif k == K_B:
                        self.mode = MODE_DRAW_BOX
                        self._etcQ.put({CROSS_CURSOR_ON:None})
                        self._updated = True
                    elif k == K_ENTER:
                        pass
                    elif k == K_ESCAPE:
                        self.mode = None


            if self._updated:
                self.put_image()
                self.put_ratio_list()
                self.put_mode()
                self._updated = False